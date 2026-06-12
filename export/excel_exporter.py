from __future__ import annotations

import os
from datetime import datetime
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from core.currency import base_currency
from ui.lang import L

if TYPE_CHECKING:
    from data.processor import InvoiceDataFrame

_GREEN = "2f8f6b"
_LIGHT_GREEN = "e8f5f0"
_RED = "e74c3c"
_ORANGE = "f39c12"
_YELLOW = "fff3cd"
_WHITE = "FFFFFF"
_DARK = "1a2332"


def _header_fill():
    return PatternFill("solid", fgColor=_GREEN)


def _header_font():
    return Font(color=_WHITE, bold=True, size=10)


def _border():
    thin = Side(style="thin", color="D0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_header_row(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()


def _autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _sheet_title(text: str) -> str:
    """Excel sheet titles: <=31 chars and no []:*?/\\ characters."""
    for ch in "[]:*?/\\":
        text = text.replace(ch, " ")
    return text[:31].strip() or "Sheet"


def export_excel(idf: "InvoiceDataFrame", output_path: str) -> str:
    """Export all invoices to a multi-sheet Excel file. Returns the output path."""
    t = L().t
    base = base_currency()
    yes, no = t("word_yes"), t("word_no")
    df = idf.get_all()
    summary = idf.get_summary()

    wb = openpyxl.Workbook()

    # ---- Sheet 1: All invoices ----
    ws_all = wb.active
    ws_all.title = _sheet_title(t("nav_invoices"))

    headers_all = [
        t("col_file"), t("col_supplier"), t("col_tax_id"), t("col_iban"),
        t("col_invoice_no"), t("col_issue"), t("col_due"), t("col_net"),
        t("col_vat"), t("col_vat_rate"), t("col_total"), t("col_currency"),
        t("col_category"), t("col_confidence"), t("col_scanned"),
        t("reason_duplicate"), t("reason_outlier"), t("reason_near_due"),
    ]
    _apply_header_row(ws_all, headers_all)
    ws_all.row_dimensions[1].height = 20

    flag_fills = {
        "duplicate": PatternFill("solid", fgColor="FFD580"),
        "outlier": PatternFill("solid", fgColor="FFBABA"),
        "near_due": PatternFill("solid", fgColor="FFF3CD"),
    }

    for r, (_, row) in enumerate(df.iterrows(), 2):
        values = [
            row["file_name"],
            row["supplier_name"],
            row["supplier_cui"],
            row["supplier_iban"],
            row["invoice_number"],
            str(row["issue_date"]) if row["issue_date"] else "",
            str(row["due_date"]) if row["due_date"] else "",
            row["subtotal"],
            row["vat_amount"],
            row["vat_rate"],
            row["total"],
            row["currency"],
            row["category"],
            f"{row['confidence_score']:.0%}",
            yes if row["is_scanned"] else no,
            yes if row["is_duplicate"] else no,
            yes if row["is_outlier"] else no,
            yes if row["is_near_due"] else no,
        ]
        fill = None
        if row["is_duplicate"]:
            fill = flag_fills["duplicate"]
        elif row["is_outlier"]:
            fill = flag_fills["outlier"]
        elif row["is_near_due"]:
            fill = flag_fills["near_due"]

        for c, val in enumerate(values, 1):
            cell = ws_all.cell(row=r, column=c, value=val)
            cell.border = _border()
            if fill:
                cell.fill = fill
            if c in (8, 9, 11):  # numeric columns
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    _autofit_columns(ws_all)

    # ---- Sheet 2: Summary ----
    ws_sum = wb.create_sheet(_sheet_title(t("report_exec_summary")))
    ws_sum["A1"] = t("xl_metric")
    ws_sum["B1"] = t("xl_value")
    ws_sum["A1"].fill = _header_fill()
    ws_sum["A1"].font = _header_font()
    ws_sum["B1"].fill = _header_fill()
    ws_sum["B1"].font = _header_font()

    rows = [
        (t("kpi_total_invoices"), summary["total_invoices"]),
        (f"{t('kpi_total_value')} ({base})", f"{summary['total_value']:,.2f}"),
        (f"{t('kpi_total_vat')} ({base})", f"{summary['total_vat']:,.2f}"),
        (t("kpi_flagged"), summary["flagged_count"]),
    ]
    for i, (k, v) in enumerate(rows, 2):
        ws_sum.cell(row=i, column=1, value=k).border = _border()
        ws_sum.cell(row=i, column=2, value=v).border = _border()

    # Per category
    row_idx = len(rows) + 3
    ws_sum.cell(row=row_idx, column=1, value=t("col_category")).fill = _header_fill()
    ws_sum.cell(row=row_idx, column=1).font = _header_font()
    ws_sum.cell(row=row_idx, column=2, value=f"{t('col_total')} ({base})").fill = _header_fill()
    ws_sum.cell(row=row_idx, column=2).font = _header_font()
    for cat, val in sorted(summary["per_category"].items(), key=lambda x: -x[1]):
        row_idx += 1
        ws_sum.cell(row=row_idx, column=1, value=cat).border = _border()
        ws_sum.cell(row=row_idx, column=2, value=round(val, 2)).border = _border()

    _autofit_columns(ws_sum)

    # ---- Sheet 3: Flagged invoices ----
    ws_flagged = wb.create_sheet(_sheet_title(t("report_flagged")))
    flagged_df = df[df["is_duplicate"] | df["is_outlier"] | df["is_near_due"]]
    _apply_header_row(ws_flagged, headers_all)
    for r, (_, row) in enumerate(flagged_df.iterrows(), 2):
        values = [
            row["file_name"], row["supplier_name"], row["supplier_cui"],
            row["supplier_iban"], row["invoice_number"],
            str(row["issue_date"]) if row["issue_date"] else "",
            str(row["due_date"]) if row["due_date"] else "",
            row["subtotal"], row["vat_amount"], row["vat_rate"],
            row["total"], row["currency"], row["category"],
            f"{row['confidence_score']:.0%}",
            yes if row["is_scanned"] else no,
            yes if row["is_duplicate"] else no,
            yes if row["is_outlier"] else no,
            yes if row["is_near_due"] else no,
        ]
        for c, val in enumerate(values, 1):
            ws_flagged.cell(row=r, column=c, value=val).border = _border()
    _autofit_columns(ws_flagged)

    wb.save(output_path)
    return output_path
